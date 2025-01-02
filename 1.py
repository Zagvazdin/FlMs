import pandas as pd
from tkinter import filedialog
from tkinter import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
# Создание корневого окна
root = Tk()
root.withdraw()  # Скрываем главное окно, чтобы не мешало

# Диалоговое окно для выбора файла
file_path = filedialog.askopenfilename()

# Загрузка данных из выбранного файла
df = pd.read_excel(file_path)

# Удаление строк с 0 по 9
df = df.drop(range(9))

# Удаление 1, 3, 4, 7, 8 столбцов
# Обратите внимание, что столбцы в Python начинаются с 0, поэтому 1-й столбец - это на самом деле столбец с индексом 0
df = df.drop(df.columns[[3, 1, 4, 7, 8]], axis=1)


new_columns = ['Клиент', 'Адрес', 'Кол-во уп', 'Сумма']
df.columns = new_columns
# Предположим, что у вас уже есть DataFrame с именем df и у вас есть столбцы 'Сумма' и 'Кол-во уп'
# Вычисляем 'Средняя цена за упаковку'
df['Средняя цена за упаковку'] = df['Сумма'] / df['Кол-во уп']

# Если в столбце 'Кол-во уп' есть нули или пропущенные значения, которые могут вызвать ошибку,
# вы можете обработать это, добавив условие проверки
df['Средняя цена за упаковку'] = df['Сумма'] / df['Кол-во уп'].where(df['Кол-во уп'] != 0, None)
df['Средняя цена за упаковку'] = df['Средняя цена за упаковку'].astype(int)

# Запрос второго файла
second_file_path = filedialog.askopenfilename()

# Загружаем данные из второго файла
df2 = pd.read_excel(second_file_path)

# Удаляем первые 10 строк и не нужные столбцы
df2 = df2.drop(range(9))
df2 = df2.drop(df2.columns[[3, 1, 4, 7, 8]], axis=1)

# Переименовываем столбцы
df2.columns = new_columns

# Вычисляем 'Средняя цена за упаковку'
df2['Средняя цена за упаковку'] = df2['Сумма'] / df2['Кол-во уп']
df2['Средняя цена за упаковку'] = df2['Средняя цена за упаковку'].astype(int)
# Объединяем данные
merged_df = df.merge(df2, on=['Клиент', 'Адрес'], suffixes=('_df1', '_df2'))

# Вычисляем разницу
merged_df['Изменение средней цены'] = merged_df['Средняя цена за упаковку_df1'] - merged_df['Средняя цена за упаковку_df2']

# Сохраняем результат
merged_df.to_excel('merged_file.xlsx', index=False)

# Выберите нужные столбцы
merged_df = merged_df[['Клиент', 'Адрес', 'Изменение средней цены']]

# Сохраните DataFrame
merged_df.to_excel('merged_file.xlsx', index=False)
# Создаем новый workbook
wb = Workbook()
ws = wb.active

# Заполняем worksheet данными из DataFrame
for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Устанавливаем ширину столбца в соответствии с содержимым
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length

# Применяем условное форматирование
greenFill = PatternFill(start_color='0000FF00',
                        end_color='0000FF00',
                        fill_type='solid')

redFill = PatternFill(start_color='00FF0000',
                      end_color='00FF0000',
                      fill_type='solid')

yellowFill = PatternFill(start_color='00FFFF00',
                         end_color='00FFFF00',
                         fill_type='solid')
column_name = 'C'
for row in ws.iter_rows(min_row=2):
    cell = ws['C' + str(row[2].row)] # 'Изменение средней цены' предположительно в третьем столбце
    value = row[2].value
    if value > 0:
        cell.fill = greenFill
    elif value < 0:
        cell.fill = redFill
    else:
        cell.fill = yellowFill
# Создание стиля границы
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
# Сохраняем файл
wb.save('merged_file.xlsx')

print(merged_df)