# styles.py
from openpyxl.styles import Border, Side, PatternFill

def get_thin_border():
    return Border(left=Side(style='thin'), 
                  right=Side(style='thin'), 
                  top=Side(style='thin'), 
                  bottom=Side(style='thin'))

def get_fill_colors():
    greenFill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
    redFill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
    yellowFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
    return greenFill, redFill, yellowFill
