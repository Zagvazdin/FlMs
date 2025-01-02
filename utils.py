import os
import pandas as pd
import numpy as np
from flask import send_file, flash, render_template, redirect, url_for
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font  # Импортируем Font для форматирования
from styles import get_thin_border, get_fill_colors  # Импортируем стили

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ['xls', 'xlsx']

def process_file(request):
    if 'od_file' not in request.files:
        return "Нет файла для загрузки", 400

    file = request.files['od_file']
    
    if not allowed_file(file.filename):
        return "Ошибка: только файлы Excel (.xls, .xlsx) разрешены.", 400

    plan = int(request.form['plan'])
    sector_number = request.form.get('sector_number')  # Получаем номер сектора
    
    if not sector_number:
        return "Ошибка: номер сектора не указан.", 400

    os.makedirs('uploads', exist_ok=True)
    os.makedirs('outputs', exist_ok=True)

    file_path = os.path.join('uploads', file.filename)
    
    try:
        file.save(file_path)
    except Exception as e:
        return f"Ошибка при сохранении файла: {str(e)}", 500

    df = pd.read_excel(file_path, engine='openpyxl')
    df = df.iloc[9:]
    df = df.drop(df.columns[[1, 3, 4, 6, 7, 8]], axis=1)
    df.columns = ['Название клиента', 'Адрес', 'Продано']
    df = df[~df['Название клиента'].isin(['ИТОГО', 'Торговых точек АКБ:'])]
    df['Продано'] = pd.to_numeric(df['Продано'], errors='coerce')
    price = df['Продано'].sum()

    if price == 0:
        return "Ошибка: Сумма продаж равна 0, проверьте данные.", 400

    k = plan / price
    df['План'] = np.round(df['Продано'] * k).astype(int)
    df.columns = ['Название клиента', 'Адрес', 'Фактическая продажа', 'План']

    output_file = os.path.join('outputs', f'Планирование Сектора_{sector_number}.xlsx')
    
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df[['Название клиента', 'Адрес', 'План']].to_excel(writer, index=False, sheet_name='Report')
            workbook = writer.book
            worksheet = writer.sheets['Report']

            worksheet.insert_rows(1)
            worksheet['A1'] = f'Отчет для сектора: {sector_number}'
            worksheet['A1'].font = Font(bold=True)

            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            thin_border = get_thin_border()
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

    except Exception as e:
        return f"Ошибка при сохранении выходного файла: {str(e)}", 500

    return send_file(output_file, as_attachment=True)

def upload_files(request, session):
    if 'username' not in session:
        flash("Вы должны войти в систему для доступа к этой странице.")
        return redirect(url_for('main.home'))

    if session.get('role') != 'admin':
        flash("У вас нет доступа к этой странице.")
        return redirect(url_for('main.home'))

    if request.method == 'POST':
        sector_number = request.form['sector_number']
        od_file = request.files['sales_file']
        output_file = request.files['planning_file']

        if not (allowed_file(od_file.filename) and allowed_file(output_file.filename)):
            return "Ошибка: только файлы Excel (.xls, .xlsx) разрешены.", 400

        df_od_full = pd.read_excel(od_file, engine='openpyxl')
        df_od = df_od_full.iloc[9:]
        df_od = df_od.drop(df_od.columns[[1, 3, 4, 6, 7, 8]], axis=1)
        df_od.columns = ['Название клиента', 'Адрес', 'Продано']
        df_od = df_od[~df_od['Название клиента'].isin(['ИТОГО', 'Торговых точек АКБ:'])]

        df_output = pd.read_excel(output_file)
        df_output.columns = ['Название клиента', 'Адрес', 'План']

        df_output['План'] = pd.to_numeric(df_output['План'], errors='coerce')
        df_od['Продано'] = pd.to_numeric(df_od['Продано'], errors='coerce')

        merged_df = pd.merge(df_od, df_output, on=['Название клиента', 'Адрес'], how='outer').fillna(0)
        merged_df['Осталось продать'] = merged_df['План'] - merged_df['Продано']
        merged_df = merged_df.drop_duplicates()

        df_odp = df_od_full.iloc[9:]
        df_odp = df_odp.drop(df_odp.columns[[1, 3, 4, 7, 8]], axis=1)
        df_odp.columns = ['Название клиента', 'Адрес', 'Продано', 'Сумма отгрузки']
        df_odp = df_odp[~df_odp['Название клиента'].isin(['ИТОГО', 'Торговых точек АКБ:'])]
        df_odp['Продано'] = pd.to_numeric(df_odp['Продано'], errors='coerce')

        final_df = pd.merge(merged_df, df_odp[['Название клиента']], on='Название клиента', how='left')
        final_df = final_df.drop_duplicates()
        final_df = final_df.sort_values(by='Осталось продать', ascending=False)

        output_file_path = f'Продажи_{sector_number}.xlsx'
        
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='Report')
            workbook = writer.book
            worksheet = writer.sheets['Report']

            worksheet.insert_rows(1)
            worksheet['A1'] = f'Отчет для сектора: {sector_number}'
            worksheet['A1'].font = Font(bold=True)

            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            thin_border = get_thin_border()
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

        return send_file(output_file_path, as_attachment=True)

    return render_template('upload.html')

def up_file(request):
    sector_number = request.form['sector_number']

    first_file = request.files['previous_month_file']
    if not allowed_file(first_file.filename):
        return "Ошибка: только файлы Excel (.xls, .xlsx) разрешены.", 400

    df = pd.read_excel(first_file).iloc[9:]
    df = df.drop(df.columns[[1, 3, 4, 7, 8]], axis=1)
    df.columns = ['Клиент', 'Адрес', 'Кол-во уп', 'Сумма']
    df['Средняя цена за упаковку'] = df['Сумма'] / df['Кол-во уп']
    df['Средняя цена за упаковку'] = df['Средняя цена за упаковку'].fillna(0).astype(int)

    second_file = request.files['current_month_file']
    if not allowed_file(second_file.filename):
        return "Ошибка: только файлы Excel (.xls, .xlsx) разрешены.", 400

    df2 = pd.read_excel(second_file).iloc[9:]
    df2 = df2.drop(df2.columns[[1, 3, 4, 7, 8]], axis=1)
    df2.columns = ['Клиент', 'Адрес', 'Кол-во уп', 'Сумма']
    df2['Средняя цена за упаковку'] = df2['Сумма'] / df2['Кол-во уп']
    df2['Средняя цена за упаковку'] = df2['Средняя цена за упаковку'].fillna(0).astype(int)

    merged_df = df.merge(df2, on=['Клиент', 'Адрес'], suffixes=('_df1', '_df2'))
    merged_df['Изменение средней цены'] = merged_df['Средняя цена за упаковку_df1'] - merged_df['Средняя цена за упаковку_df2']
    
    merged_df = merged_df[['Клиент', 'Адрес', 'Изменение средней цены']]

    wb = Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    greenFill, redFill, yellowFill = get_fill_colors()
    
    for row in ws.iter_rows(min_row=2):
        cell = row[2]
        value = cell.value
        if value > 0:
            cell.fill = greenFill
        elif value < 0:
            cell.fill = redFill
        else:
            cell.fill = yellowFill

    ws.insert_rows(1)
    ws['A1'] = f'Отчет для сектора: {sector_number}'
    ws['A1'].font = Font(bold=True)

    output_file_path = f'Изменение_цен_{sector_number}.xlsx'
    wb.save(output_file_path)

    return send_file(output_file_path, as_attachment=True)
